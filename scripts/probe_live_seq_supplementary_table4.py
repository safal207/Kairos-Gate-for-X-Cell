#!/usr/bin/env python3
"""Probe Live-seq Supplementary Table 4 without treating it as replication."""

from __future__ import annotations

import hashlib
import json
import math
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def numeric(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def main(argv: list[str]) -> int:
    if len(argv) != 3:
        print("usage: probe_live_seq_supplementary_table4.py INPUT.xlsx OUTPUT.json", file=sys.stderr)
        return 2

    source = Path(argv[1])
    output = Path(argv[2])
    raw = source.read_bytes()
    workbook = load_workbook(source, read_only=True, data_only=False)

    report: dict[str, Any] = {
        "source_file": source.name,
        "sha256": hashlib.sha256(raw).hexdigest(),
        "size_bytes": len(raw),
        "sheet_count": len(workbook.sheetnames),
        "sheets": [],
        "nfkbia_hits": [],
        "model_ranking_summary": None,
        "interpretation_boundary": {
            "artifact_role": "original-study model-ranking supplement",
            "external_replication": False,
            "causal_identification": False,
            "notes": "The workbook can describe the original ranking and multiplicity behavior. It cannot establish biological independence, direct replication, or causality."
        }
    }

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        preview_rows: list[list[str]] = []
        nonempty = 0
        max_seen_columns = 0
        matrix: list[list[Any]] = []

        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            raw_values = list(row)
            values = [text(value) for value in raw_values]
            while values and values[-1] == "":
                values.pop()
                raw_values.pop()
            if not values:
                continue

            matrix.append(raw_values)
            nonempty += 1
            max_seen_columns = max(max_seen_columns, len(values))
            if len(preview_rows) < 8:
                preview_rows.append(values[:20])

            for column_index, value in enumerate(values, start=1):
                if "nfkbia" in value.lower():
                    report["nfkbia_hits"].append({
                        "sheet": sheet_name,
                        "row": row_index,
                        "column": column_index,
                        "value": value,
                        "row_preview": values[:20]
                    })

        report["sheets"].append({
            "name": sheet_name,
            "max_row_declared": ws.max_row,
            "max_column_declared": ws.max_column,
            "nonempty_rows": nonempty,
            "max_seen_columns": max_seen_columns,
            "first_nonempty_rows": preview_rows
        })

        if not matrix:
            continue
        headers = [text(value) for value in matrix[0]]
        required = {"pval_lm", "fdr_lm", "r2", "coef", "gene", "symbol", "pval_bootstrap", "fdr_bootstrap"}
        if not required.issubset(set(headers)):
            continue

        index = {name: headers.index(name) for name in required}
        genes: list[dict[str, Any]] = []
        for source_row, row in enumerate(matrix[1:], start=2):
            symbol = text(row[index["symbol"]] if index["symbol"] < len(row) else None)
            if not symbol:
                continue
            genes.append({
                "source_row": source_row,
                "gene": text(row[index["gene"]] if index["gene"] < len(row) else None),
                "symbol": symbol,
                "pval_lm": numeric(row[index["pval_lm"]] if index["pval_lm"] < len(row) else None),
                "fdr_lm": numeric(row[index["fdr_lm"]] if index["fdr_lm"] < len(row) else None),
                "r2": numeric(row[index["r2"]] if index["r2"] < len(row) else None),
                "coef": numeric(row[index["coef"]] if index["coef"] < len(row) else None),
                "pval_bootstrap": numeric(row[index["pval_bootstrap"]] if index["pval_bootstrap"] < len(row) else None),
                "fdr_bootstrap": numeric(row[index["fdr_bootstrap"]] if index["fdr_bootstrap"] < len(row) else None),
            })

        ranked = sorted(genes, key=lambda item: (float("inf") if item["pval_lm"] is None else item["pval_lm"]))
        nfkbia = next((item for item in ranked if item["symbol"].lower() == "nfkbia"), None)
        if nfkbia is None:
            continue
        rank = ranked.index(nfkbia) + 1
        next_gene = ranked[rank] if rank < len(ranked) else None
        r2_gap = None
        if next_gene and nfkbia["r2"] is not None and next_gene["r2"] is not None:
            r2_gap = nfkbia["r2"] - next_gene["r2"]

        def count_at(field: str, threshold: float) -> int:
            return sum(1 for item in genes if item[field] is not None and item[field] <= threshold)

        report["model_ranking_summary"] = {
            "sheet": sheet_name,
            "tested_gene_count": len(genes),
            "nfkbia_rank_by_lm_pvalue": rank,
            "nfkbia": nfkbia,
            "next_ranked_gene": next_gene,
            "r2_gap_to_next_ranked_gene": r2_gap,
            "lm_fdr_counts": {
                "le_0_05": count_at("fdr_lm", 0.05),
                "le_0_10": count_at("fdr_lm", 0.10),
                "le_0_20": count_at("fdr_lm", 0.20)
            },
            "bootstrap_fdr_counts": {
                "le_0_05": count_at("fdr_bootstrap", 0.05),
                "le_0_10": count_at("fdr_bootstrap", 0.10),
                "le_0_20": count_at("fdr_bootstrap", 0.20)
            },
            "top_10_by_lm_pvalue": ranked[:10],
            "stability_interpretation": {
                "lm_ranking": "Nfkbia is the top nominal linear-model feature in this workbook.",
                "bootstrap_multiplicity": "Bootstrap FDR must be inspected separately; a weak or null corrected result limits claims that the top feature is uniquely stable.",
                "winner_curse_risk": "high when feature selection and effect estimation use the same small target sample without direct external replication"
            }
        }

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    summary = report.get("model_ranking_summary") or {}
    print(json.dumps({
        "sha256": report["sha256"],
        "size_bytes": report["size_bytes"],
        "sheet_count": report["sheet_count"],
        "nfkbia_hit_count": len(report["nfkbia_hits"]),
        "tested_gene_count": summary.get("tested_gene_count"),
        "nfkbia_rank": summary.get("nfkbia_rank_by_lm_pvalue"),
        "nfkbia_lm_fdr": (summary.get("nfkbia") or {}).get("fdr_lm"),
        "nfkbia_bootstrap_fdr": (summary.get("nfkbia") or {}).get("fdr_bootstrap"),
        "output": str(output)
    }, indent=2))

    if report["sheet_count"] < 1:
        print("workbook contains no sheets", file=sys.stderr)
        return 1
    if len(report["nfkbia_hits"]) < 1:
        print("Nfkbia was not found in Supplementary Table 4", file=sys.stderr)
        return 1
    if report["model_ranking_summary"] is None:
        print("expected model-ranking columns were not found", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
