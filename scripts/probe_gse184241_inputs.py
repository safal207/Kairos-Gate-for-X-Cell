#!/usr/bin/env python3
"""Probe official GSE184241 processed inputs without making scientific claims."""

from __future__ import annotations

import argparse
import gzip
import hashlib
import json
import shlex
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def split_geo_line(line: str) -> list[str]:
    """Split the quoted whitespace-delimited table emitted by GEO."""
    return shlex.split(line.rstrip("\n\r"), posix=True)


def probe_counts(path: Path) -> dict[str, Any]:
    with gzip.open(path, "rt", encoding="utf-8", errors="replace") as handle:
        header = split_geo_line(handle.readline())
        preview: list[list[str]] = []
        row_count = 0
        nonzero_lengths: list[int] = []
        observed_widths: set[int] = set()
        for line in handle:
            row_count += 1
            values = split_geo_line(line)
            observed_widths.add(len(values))
            if len(preview) < 8:
                preview.append(values[: min(12, len(values))])
            if row_count <= 100:
                nonzero_lengths.append(sum(value not in {"", "0", "0.0"} for value in values[1:]))

    return {
        "delimiter": "quoted_whitespace",
        "row_count_excluding_header": row_count,
        "cell_column_count": len(header),
        "data_row_widths_observed": sorted(observed_widths),
        "first_20_cell_names": header[:20],
        "last_10_cell_names": header[-10:],
        "first_8_rows_first_12_values": preview,
        "first_100_genes_nonzero_cell_count_summary": {
            "min": min(nonzero_lengths) if nonzero_lengths else None,
            "max": max(nonzero_lengths) if nonzero_lengths else None,
        },
    }


def probe_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        preview: list[list[Any]] = []
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 25), values_only=True):
            preview.append([value for value in row[: min(sheet.max_column, 20)]])
        sheets.append(
            {
                "title": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "first_25_rows_first_20_values": preview,
            }
        )
    return {"sheet_count": len(sheets), "sheets": sheets}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--counts", required=True)
    parser.add_argument("--barcodes", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--source-revision", required=True)
    args = parser.parse_args()

    counts = Path(args.counts)
    barcodes = Path(args.barcodes)
    output = Path(args.output)

    result = {
        "schema_version": "0.2.0",
        "artifact_type": "gse184241_input_probe",
        "source_revision": args.source_revision,
        "dataset": {
            "accession": "GSE184241",
            "organism": "Homo sapiens",
            "declared_biological_group": "donor",
            "declared_donors": ["Donor1", "Donor2", "Donor3"],
            "direct_same_cell_future_response": False,
        },
        "files": {
            "counts": {
                "filename": counts.name,
                "sha256": sha256_file(counts),
                "size_bytes": counts.stat().st_size,
                "probe": probe_counts(counts),
            },
            "barcodes": {
                "filename": barcodes.name,
                "sha256": sha256_file(barcodes),
                "size_bytes": barcodes.stat().st_size,
                "probe": probe_workbook(barcodes),
            },
        },
        "claim_boundary": {
            "data_structure_only": True,
            "benchmark_result_generated": False,
            "geneformer_inference_executed": False,
            "same_cell_prediction_established": False,
        },
    }

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(result, indent=2, sort_keys=True, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"output": str(output), "counts_sha256": result["files"]["counts"]["sha256"], "barcodes_sha256": result["files"]["barcodes"]["sha256"]}, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
